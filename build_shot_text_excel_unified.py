import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import os
import re
import threading
from collections import Counter
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


ROOT = Path(os.environ.get("SHOT_CUTTING_PROJECT_ROOT", str(Path.cwd())))
OCR_CACHE_VERSION = "ocr_crop720_v1"
OCR_MAX_WIDTH = 720
_OCR_LOCAL = threading.local()

FIELDS = [
    "视频链接",
    "镜头",
    "时间",
    "视频结构（画面）",
    "文案框架",
    "文案",
    "用户视角",
    "配音",
    "音效39个",
    "视频亮点",
]


def imread(path: Path) -> np.ndarray | None:
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def read_video_frame(cap: cv2.VideoCapture, frame_no: int) -> np.ndarray | None:
    cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, int(frame_no)))
    ok, frame = cap.read()
    return frame if ok else None


def read_video_frames(video_path: Path, frame_numbers: list[int] | set[int]) -> dict[int, np.ndarray]:
    wanted = sorted({max(0, int(frame_no)) for frame_no in frame_numbers})
    if not wanted:
        return {}
    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        return {}

    frames: dict[int, np.ndarray] = {}
    if len(wanted) >= 180:
        wanted_set = set(wanted)
        frame_no = 0
        while wanted_set:
            ok, frame = cap.read()
            if not ok:
                break
            if frame_no in wanted_set:
                frames[frame_no] = frame.copy()
                wanted_set.remove(frame_no)
            frame_no += 1
    else:
        for frame_no in wanted:
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_no)
            ok, frame = cap.read()
            if ok:
                frames[frame_no] = frame.copy()
    cap.release()
    return frames


def normalize_subtitle(text: str) -> str:
    text = re.sub(r"\s+", "", text.strip())
    text = text.replace("眞", "真")
    text = text.replace("不換", "不换")
    text = text.replace("將就", "将就")
    text = text.replace("由干", "由于")
    text = text.replace("做出回", "做出口")
    text = text.replace("再不围", "再不买")
    text = text.replace("备受青的", "备受青睐的")
    text = text.replace("终干", "终于")
    text = text.replace("查接", "直接")
    text = text.replace("道接", "直接")
    text = text.replace("档部", "裆部")
    text = text.replace("y宇", "y字")
    text = text.replace("Y宇", "Y字")
    text = text.replace("细带", "绷带")
    text = text.replace("烂裤权", "烂裤衩")
    text = text.replace("折开", "拆开")
    text = text.replace("自已", "自己")
    text = text.replace("呢么少人", "这么少人")
    text = text.replace("昨不穿", "咋不穿")
    text = text.replace("底档", "底裆")
    text = text.replace("小底裆", "小底裆")
    text = text.replace("优feel", "ufeel")
    text = text.replace("开俊即穿", "开袋即穿")
    text = text.replace("买76条", "买了6条")
    text = text.replace("开装即宇", "开袋即穿")
    text = re.sub(r".*美宅私物[（(]?开[袋业装][^/／]*[/／]", "", text)
    text = text.replace("问了女知道", "问了才知道")
    text = text.replace("问了知道", "问了才知道")
    text = text.replace("具的很明显", "真的很明显")
    text = text.replace("属干", "属于")
    text = text.replace("相当干", "相当于")
    text = text.replace("本白色底裆", "这个白色底裆")
    text = text.replace("洗7", "洗了")
    text = text.replace("身林", "身材")
    text = text.replace("身树", "身材")
    text = text.replace("不知道他肚子越大", "不知道它肚子越大")
    text = text.replace("不知道她肚子越大", "不知道它肚子越大")
    text = re.sub(r"^草本初色/", "", text)
    if "草本初色Herb" in text:
        text = "草本初色"
    if any(
        token in text
        for token in [
            "日蚕丝粉底液内裤",
            "白野蚕丝粉底液内裤",
            "山茶花香氛肌底裤",
            "塑形裤|呼吸",
            "美宅私物开袋即穿",
            "私物开袋即穿",
            "物开袋即穿",
        ]
    ):
        text = ""
    replacements = {
        "苹身": "半身",
        "善通": "普通",
        "普逼": "普通",
        "内认": "内衣",
        "不间": "不闷",
        "两杆小风扇": "两个小风扇",
        "香诗": "嫦香诗",
        "小黑磁": "小黑瓷",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    if text == "文胸":
        text = "内衣"
    return text


def looks_like_subtitle(text: str, score: float) -> bool:
    if score < 0.62:
        return False
    text = normalize_subtitle(text)
    if len(text) < 4:
        return False
    if re.fullmatch(r"[\d:：./ -]+", text):
        return False
    if not re.search(r"[\u4e00-\u9fff]", text):
        return False
    return text not in {"7:00", "jMo", "MO", "No", "mO"}


def subtitle_crop_specs(mode: str) -> list[tuple[str, float, float, float, float]]:
    if mode == "auto":
        mode = "bottom"
    if mode in {"middle", "center"}:
        mode = "wide"
    if mode == "top":
        return [("top", 0.06, 0.46, 0.00, 0.38)]
    if mode == "wide":
        return [("wide", 0.15, 0.88, 0.00, 1.00)]
    if mode == "top-bottom":
        return [
            ("top", 0.06, 0.46, 0.00, 0.38),
            ("bottom", 0.55, 0.86, 0.38, 0.92),
        ]
    return [("bottom", 0.45, 0.92, 0.22, 0.95)]


def looks_like_white_subtitle_region(crop: np.ndarray, box: list[list[float]]) -> bool:
    ys = [point[1] for point in box]
    xs = [point[0] for point in box]
    x0 = max(0, int(min(xs)) - 4)
    x1 = min(crop.shape[1], int(max(xs)) + 4)
    y0 = max(0, int(min(ys)) - 4)
    y1 = min(crop.shape[0], int(max(ys)) + 4)
    roi = crop[y0:y1, x0:x1]
    if roi.size == 0:
        return False
    bgr = roi.reshape(-1, 3)
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV).reshape(-1, 3)
    white_ratio = float(np.mean((bgr[:, 0] > 170) & (bgr[:, 1] > 170) & (bgr[:, 2] > 170)))
    saturated_ratio = float(np.mean((hsv[:, 1] > 70) & (hsv[:, 2] > 120)))
    orange_ratio = float(np.mean((bgr[:, 2] > 150) & (bgr[:, 1] > 80) & (bgr[:, 0] < 150) & (hsv[:, 1] > 50)))
    if white_ratio < 0.025:
        return False
    if saturated_ratio > 0.35 and orange_ratio > 0.12:
        return False
    return True


def resize_crop_for_ocr(crop: np.ndarray) -> tuple[np.ndarray, float]:
    h, w = crop.shape[:2]
    if w <= OCR_MAX_WIDTH:
        return crop, 1.0
    scale = OCR_MAX_WIDTH / float(w)
    resized = cv2.resize(crop, (OCR_MAX_WIDTH, max(1, int(h * scale))), interpolation=cv2.INTER_AREA)
    return resized, 1.0 / scale


def scale_box(box: list[list[float]], scale_back: float) -> list[list[float]]:
    if scale_back == 1.0:
        return box
    return [[float(point[0]) * scale_back, float(point[1]) * scale_back] for point in box]


def ocr_subtitle(ocr: RapidOCR, frame: np.ndarray | None, mode: str = "bottom") -> str:
    if frame is None:
        return ""
    h, w = frame.shape[:2]
    candidates: list[tuple[float, float, str]] = []
    for _, y0, y1, min_center, max_center in subtitle_crop_specs(mode):
        crop_top = int(h * y0)
        crop = frame[crop_top : int(h * y1), 0:w]
        ocr_crop, scale_back = resize_crop_for_ocr(crop)
        result, _ = ocr(ocr_crop)
        if not result:
            continue

        crop_h = max(1, crop.shape[0])
        for item in result:
            box, raw_text, score = scale_box(item[0], scale_back), str(item[1]), float(item[2])
            text = normalize_subtitle(raw_text)
            if not looks_like_subtitle(text, score):
                continue
            ys = [point[1] for point in box]
            xs = [point[0] for point in box]
            center_y = sum(ys) / len(ys)
            box_h = max(ys) - min(ys)
            box_w = max(xs) - min(xs)
            if center_y < crop_h * min_center or center_y > crop_h * max_center:
                continue
            if box_h < 20 or box_w < w * 0.18:
                continue
            if not looks_like_white_subtitle_region(crop, box):
                continue
            candidates.append((crop_top + center_y, min(xs), text))

    if not candidates:
        return ""
    candidates.sort()
    merged: list[str] = []
    for _, _, text in candidates:
        if text not in merged:
            merged.append(text)
    if merged:
        brand_noise = {"草本初色", "梵客隆V", "梵客隆Y", "梵客隆v", "梵客隆y", "桥客隆V"}
        merged = [
            text
            for text in merged
            if text not in brand_noise
            and not any(token in text for token in ["Herbfree", "Herbfee", "Herb", "PIER", "N°5"])
        ]
    return " / ".join(merged)


def choose_subtitle_region(
    shots: list[dict[str, Any]],
    video_output_dir: Path,
    requested_region: str,
    sample_size: int = 28,
) -> str:
    if requested_region != "auto":
        if requested_region in {"middle", "center"}:
            return "wide"
        return requested_region
    if not shots:
        return "bottom"

    sample_indices = sorted(
        {
            0,
            min(len(shots) - 1, 1),
            min(len(shots) - 1, 2),
            *[int(round(i * (len(shots) - 1) / max(1, sample_size - 1))) for i in range(sample_size)],
        }
    )
    ocr = RapidOCR()
    bottom_hits = 0
    wide_hits = 0
    wide_only = 0
    early_wide_only = 0

    for index in sample_indices:
        shot = shots[index]
        frame = imread(resolve_evidence_path(video_output_dir, shot))
        bottom_text = ocr_subtitle(ocr, frame, "bottom")
        if bottom_text:
            bottom_hits += 1
            continue
        wide_text = ocr_subtitle(ocr, frame, "wide")
        if wide_text:
            wide_hits += 1
            wide_only += 1
            if float(shot.get("start_time_ms", 0.0)) <= 30000.0:
                early_wide_only += 1

    if wide_only >= 3 or early_wide_only >= 2 or wide_hits > bottom_hits + 1:
        return "wide"
    return "bottom"


def extract_shot_subtitle(
    ocr: RapidOCR,
    cap: cv2.VideoCapture,
    shot: dict[str, Any],
    subtitle_region: str = "bottom",
    evidence_text: str | None = None,
    full_sampling: bool = False,
) -> str:
    start = int(shot.get("start_frame", 0))
    end = int(shot.get("end_frame", start))
    rep = int(shot.get("representative_frame", start))

    if evidence_text is None:
        evidence = imread(Path(str(shot.get("evidence_frame", ""))))
        evidence_text = ocr_subtitle(ocr, evidence, subtitle_region)
    if evidence_text:
        return evidence_text

    rep_text = ocr_subtitle(ocr, read_video_frame(cap, rep), subtitle_region)
    if rep_text:
        return rep_text
    if not full_sampling:
        return ""

    sample_frames = [
        start,
        start + max(0, (end - start) // 2),
        max(start, end - 2),
    ]
    texts: list[str] = []
    for frame_no in dict.fromkeys(sample_frames):
        text = ocr_subtitle(ocr, read_video_frame(cap, frame_no), subtitle_region)
        if text:
            texts.append(text)
    if not texts:
        return ""

    counts = Counter(texts)
    return max(counts, key=lambda item: (counts[item], len(item)))


def make_thumb(image_path: Path, shot_id: str, thumb_dir: Path) -> Path:
    output = thumb_dir / f"{shot_id}.jpg"
    with Image.open(image_path) as img:
        img.thumbnail((190, 338), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (190, 338), "white")
        x = (190 - img.width) // 2
        y = (338 - img.height) // 2
        canvas.paste(img.convert("RGB"), (x, y))
        canvas.save(output, quality=88)
    return output


def split_transcript(text: str) -> list[str]:
    text = re.sub(r"\s+", "", text or "")
    if not text:
        return []
    pieces = re.split(r"(?<=[。！？!?])", text)
    units = [piece.strip() for piece in pieces if piece.strip()]
    if len(units) <= 1:
        units = [piece.strip() for piece in re.split(r"[，,。！？!?；;]", text) if piece.strip()]
    return units


def transcript_copy_for_index(index: int, total: int, units: list[str], previous: str) -> str:
    if not units:
        return ""
    unit_index = min(len(units) - 1, int((index - 1) * len(units) / max(1, total)))
    text = units[unit_index]
    return "" if text == previous else text


def load_timed_transcript(path: Path | None) -> list[dict[str, Any]]:
    if not path:
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    raw_segments = data.get("segments", data) if isinstance(data, dict) else data
    segments: list[dict[str, Any]] = []
    for item in raw_segments or []:
        if not isinstance(item, dict):
            continue
        text = normalize_subtitle(str(item.get("text", "")))
        if not text:
            continue
        start_ms = item.get("start_ms")
        end_ms = item.get("end_ms")
        if start_ms is None:
            start_ms = float(item.get("start", 0.0)) * 1000.0
        if end_ms is None:
            end_ms = float(item.get("end", item.get("start", 0.0))) * 1000.0
        segments.append({"start_ms": float(start_ms), "end_ms": float(end_ms), "text": text})
    segments.sort(key=lambda item: (item["start_ms"], item["end_ms"]))
    return segments


def timed_transcript_copy_for_shot(shot: dict[str, Any], segments: list[dict[str, Any]]) -> str:
    if not segments:
        return ""
    start_ms = float(shot.get("start_time_ms", 0.0))
    end_ms = float(shot.get("end_time_ms", start_ms))
    min_overlap_ms = 80.0
    texts: list[str] = []
    for segment in segments:
        seg_start = float(segment["start_ms"])
        seg_end = max(seg_start, float(segment["end_ms"]))
        if seg_end < start_ms - 120.0:
            continue
        if seg_start > end_ms + 120.0:
            break
        overlap = min(end_ms, seg_end) - max(start_ms, seg_start)
        starts_inside = start_ms <= seg_start <= end_ms
        ends_inside = start_ms <= seg_end <= end_ms
        if overlap >= min_overlap_ms or starts_inside or ends_inside:
            text = str(segment["text"]).strip()
            if text and text not in texts:
                texts.append(text)
    return " / ".join(texts)


def classify_plan(
    index: int,
    shot: dict[str, Any],
    copy: str,
    repeated_copy: bool,
    product_hint: str = "",
) -> dict[str, str]:
    text = copy
    notes = str(shot.get("notes", ""))
    pace = str(shot.get("pace_tag", ""))
    start_ms = float(shot.get("start_time_ms", 0.0))

    visual = "模特/产品画面承接"
    framework = ""
    viewer = ""

    if index <= 3:
        visual = "开场舒适痛点"
        framework = "舒适需求开场"
        viewer = "用舒适或换新需求把目标用户带入，建立继续看的理由。"
    elif any(word in text for word in ["我老婆只穿", "翻个身"]):
        visual = "强反差吸睛画面"
        framework = "反常识钩子"
        viewer = "被反差说法吸引，想知道为什么这条内裤值得买。"
    elif any(word in text for word in ["顶得住", "算我输", "哈喽大家好"]):
        visual = "口播字幕画面"
        framework = "悬念承接"
        viewer = "口播继续放大好奇心，等待产品理由展开。"
    elif any(word in text for word in ["福利", "买一发三", "价格", "好几十"]):
        visual = "活动利益点画面"
        framework = "活动利益点"
        viewer = "价格对比和福利明确，容易产生现在入手的想法。"
    elif "普通内衣" in text or "不要再穿" in text:
        visual = "普通内衣痛点对比"
        framework = "痛点提醒"
        viewer = "用户会联想到普通内衣勒、空杯、跑杯或不舒服的问题。"
    elif any(word in text for word in ["提拉内衣", "聚拢内衣", "都市丽人", "嫦香诗", "香诗"]):
        visual = "产品正式露出"
        framework = "产品给到解决方案"
        viewer = "明确产品类型和品牌，开始判断是否适合自己。"
    elif any(word in text for word in ["塑身", "美嘉挺", "婷美", "美形", "塑形", "收腹裤", "收腹内裤", "塑身内衣", "提臀裤", "全包臀"]) or (
        any(word in product_hint for word in ["塑身", "美嘉挺", "婷美", "美形"])
        and any(word in text for word in ["内衣", "内裤", "收腹", "肚子", "臀", "上身", "腰", "穿"])
    ):
        visual = "塑身/收腹内裤产品露出"
        framework = "产品给到解决方案"
        viewer = "明确产品是塑身收腹类内裤，开始判断收腹、提臀、无痕和穿搭改善效果。"
    elif any(word in text for word in ["一口气", "4件", "两件", "囤"]):
        visual = "多件囤货/复购证明"
        framework = "数量价值证明"
        viewer = "看到一次买多件，会觉得产品有换新和囤货价值。"
    elif any(word in text for word in ["内裤", "蕾丝", "底档", "裆", "小花园"]):
        visual = "女士内裤产品露出"
        framework = "产品给到解决方案"
        viewer = "明确产品是女士内裤，开始判断舒适、卫生、无痕和穿搭适配。"
    elif any(word in text for word in ["收腹", "肚子", "妈妈臀", "臀部肉肉", "全包臀", "上身"]):
        visual = "上身效果展示"
        framework = "身材痛点+效果承诺"
        viewer = "联想到自己的身材和穿搭尴尬，关注上身改善效果。"
    elif any(word in text for word in ["旧内衣", "全换", "全扔", "换成"]):
        visual = "旧内衣换新痛点"
        framework = "换新理由强化"
        viewer = "旧内衣被替换，强化用户也该升级内衣的心理。"
    elif any(word in text for word in ["越小的胸", "越散的胸", "下垂", "胸型", "归位", "垂不垂"]):
        visual = "聚拢胸型改善"
        framework = "胸型痛点解决"
        viewer = "小胸、散胸、下垂胸用户会关注是否能聚拢、承托并改善胸型。"
    elif any(word in text for word in ["外貌加分", "好看了", "显瘦", "蜜桃胸", "蚂蚁腰"]):
        visual = "上身效果强化"
        framework = "结果型卖点"
        viewer = "把内衣效果和外貌加分、身材变好直接关联，容易形成购买想象。"
    elif any(word in text for word in ["春夏", "穿啥都好看", "白t", "白T", "罩衫"]):
        visual = "春夏穿搭场景"
        framework = "穿搭场景验证"
        viewer = "把产品放进白T、罩衫、薄衣服等真实场景里判断实用性。"
    elif any(word in text for word in ["娜扎", "同款"]):
        visual = "明星同款背书"
        framework = "种草信任背书"
        viewer = "明星同款会降低陌生感，提高继续看的兴趣。"
    elif any(word in text for word in ["27年", "4500", "6300万", "国民", "门店", "女性选择", "国货品牌", "线下专柜"]):
        visual = "品牌实力背书"
        framework = "品牌信任建立"
        viewer = "通过品牌年限、门店和用户规模判断是否可靠。"
    elif any(word in text for word in ["回馈老客户", "老客户", "专柜老贵"]):
        visual = "品牌活动信任"
        framework = "活动合理化"
        viewer = "用线下专柜价格和老客户回馈解释优惠，降低低价疑虑。"
    elif any(word in text for word in ["喜马拉雅", "太极磁", "四道工序", "轻氧SPA", "软又滑"]):
        visual = "面料科技/触感证明"
        framework = "材料卖点拆解"
        viewer = "从面料来源、工艺和触感判断是否真的舒服。"
    elif any(word in text for word in ["轻盈", "透气", "舒适", "裸感"]):
        visual = "轻盈舒适证明"
        framework = "舒适卖点"
        viewer = "关注内衣是否闷、勒、厚，以及春夏穿着是否舒服。"
    elif any(word in text for word in ["不闷", "透气网孔", "凉感", "小风扇", "双重透气"]):
        visual = "透气凉感证明"
        framework = "夏季舒适卖点"
        viewer = "夏天用户最担心闷热出汗，这里用网孔和凉感降低顾虑。"
    elif any(word in text for word in ["杯垫", "防凸", "跑杯"]):
        visual = "杯垫/防凸点证明"
        framework = "结构痛点解决"
        viewer = "解决杯垫移位、凸点和清洗后跑杯的顾虑。"
    elif any(word in text for word in ["高定无痕", "无痕面料", "无痕", "留印"]):
        visual = "无痕外穿证明"
        framework = "尴尬痛点解决"
        viewer = "关注薄衣服外穿是否会露边、勒痕或尴尬留印。"
    elif any(word in text for word in ["白薄透", "薄裤子", "安全裤", "不露", "露印", "勒痕"]):
        visual = "内裤无痕穿搭证明"
        framework = "尴尬痛点解决"
        viewer = "关注白色、薄透、紧身衣物下是否会露出内裤痕迹。"
    elif any(word in text for word in ["底档加长", "底档加宽", "磨", "卫生环境", "舒道", "下半身"]):
        visual = "底档舒适卫生证明"
        framework = "舒适卫生卖点"
        viewer = "关注底档长度、摩擦感和女性私处卫生舒适。"
    elif any(word in text for word in ["柔软", "亲肤", "透气", "新疆棉", "棉花香", "软糯"]):
        visual = "内裤面料触感证明"
        framework = "面料舒适卖点"
        viewer = "从柔软、亲肤、透气判断是否适合日常贴身穿。"
    elif any(word in text for word in ["镂空", "蕾丝", "面料", "无痕", "紧身裤", "揉搓"]):
        visual = "蕾丝面料细节"
        framework = "产品核心卖点"
        viewer = "既想要好看，也在意贴身穿是否舒服、不勒、不显痕。"
    elif any(word in text for word in ["独立包装", "密封包装", "二手", "包装"]):
        visual = "包装信任证明"
        framework = "打消顾虑"
        viewer = "独立密封包装能降低卫生和二手内裤顾虑。"
    elif any(word in text for word in ["健身", "运动"]):
        visual = "运动穿着场景"
        framework = "穿搭场景验证"
        viewer = "代入运动或紧身穿搭场景，判断是否适合自己。"
    elif any(word in text for word in ["副乳", "侧收", "支撑", "腋下", "双c位", "双C位", "软支撑", "稳稳承托"]):
        visual = "侧收软支撑证明"
        framework = "支撑结构拆解"
        viewer = "关注副乳收拢、侧边稳定和无钢圈支撑是否够用。"
    elif any(word in text for word in ["美背", "交叉", "露背", "肩带", "深V"]):
        visual = "美背/肩带穿法展示"
        framework = "穿搭场景验证"
        viewer = "判断吊带、露背、薄上衣等场景能不能搭配。"
    elif any(word in text for word in ["无束缚", "不压胸", "没有钢圈", "无钢圈", "舒服"]):
        visual = "无钢圈舒适证明"
        framework = "舒适痛点解决"
        viewer = "关注聚拢承托的同时是否会勒、压胸或有钢圈束缚感。"
    elif any(word in text for word in ["活动", "价格", "拍", "下单", "闭眼", "两件只要", "有货"]):
        visual = "活动转化收口"
        framework = "价格活动促单"
        viewer = "卖点建立后，价格和活动推动用户下单。"
    elif is_fast_proof_shot(shot):
        visual = "快速单帧证明"
        framework = "快切卖点补充"
        viewer = "同场景展示文字变化或同产品颜色属性变化，适合单独保留做证明。"
    elif start_ms < 12000:
        visual = "开场产品/上身切换"
        framework = "开场钩子承接"
        viewer = "用连续快切把痛点、产品和上身状态串起来。"

    if repeated_copy:
        framework = ""
        viewer = ""

    return {
        "visual": visual,
        "framework": framework,
        "copy": text,
        "viewer": viewer,
        "voice": "生活化种草口播，语速轻快，重点词稍加强。" if framework else "",
        "sound": "轻快BGM / 字幕pop / 细节ding" if framework else "",
        "highlight": "该镜头承担卖点证明或转场承接作用。" if framework else "",
    }


def resolve_report(video_output_dir: Path, report_mode: str) -> tuple[Path, dict[str, Any]]:
    report_sets = {
        "calibrated": [video_output_dir / "model_calibrated" / "optimized_shot_report.json"],
        "reference": [video_output_dir / "reference_optimized" / "optimized_shot_report.json"],
        "model": [video_output_dir / "model_optimized" / "model_optimized_shot_report.json"],
        "auto": [
            video_output_dir / "model_calibrated" / "optimized_shot_report.json",
            video_output_dir / "reference_optimized" / "optimized_shot_report.json",
            video_output_dir / "model_optimized" / "model_optimized_shot_report.json",
        ],
    }
    candidates = report_sets[report_mode]
    for path in candidates:
        if path.exists():
            return path, json.loads(path.read_text(encoding="utf-8"))
    raise FileNotFoundError(f"No shot report found under {video_output_dir}")


def resolve_evidence_path(video_output_dir: Path, shot: dict[str, Any]) -> Path:
    path = Path(str(shot.get("evidence_frame", "")))
    if path.exists():
        return path
    return video_output_dir / "model_optimized" / "evidence" / path.name


def get_thread_ocr() -> RapidOCR:
    ocr = getattr(_OCR_LOCAL, "ocr", None)
    if ocr is None:
        ocr = RapidOCR()
        _OCR_LOCAL.ocr = ocr
    return ocr


def ocr_cache_key(path: Path, subtitle_region: str) -> str:
    stat = path.stat()
    return f"{OCR_CACHE_VERSION}|{subtitle_region}|{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}"


def video_frame_ocr_cache_key(video_path: Path, frame_no: int, subtitle_region: str) -> str:
    stat = video_path.stat()
    return (
        f"{OCR_CACHE_VERSION}|{subtitle_region}|{video_path.resolve()}|"
        f"{stat.st_size}|{stat.st_mtime_ns}|frame:{int(frame_no)}"
    )


def load_ocr_cache(cache_path: Path | None) -> dict[str, str]:
    if not cache_path or not cache_path.exists():
        return {}
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def save_ocr_cache(cache_path: Path | None, cache: dict[str, str]) -> None:
    if not cache_path:
        return
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def ocr_evidence_file(path: Path, subtitle_region: str) -> str:
    return ocr_subtitle(get_thread_ocr(), imread(path), subtitle_region)


def ocr_frame_image(frame: np.ndarray, subtitle_region: str) -> str:
    return ocr_subtitle(get_thread_ocr(), frame, subtitle_region)


def precompute_evidence_subtitles(
    shots: list[dict[str, Any]],
    video_output_dir: Path,
    subtitle_region: str,
    workers: int,
    cache_path: Path | None,
) -> dict[str, str]:
    cache = load_ocr_cache(cache_path)
    subtitles: dict[str, str] = {}
    pending: list[tuple[str, Path, str]] = []

    for shot in shots:
        evidence_path = resolve_evidence_path(video_output_dir, shot)
        key = ocr_cache_key(evidence_path, subtitle_region)
        shot_id = str(shot.get("shot_id", ""))
        if key in cache:
            subtitles[shot_id] = normalize_subtitle(str(cache[key]))
        else:
            pending.append((shot_id, evidence_path, key))

    if not pending:
        return subtitles

    workers = max(1, int(workers or 1))
    if workers == 1:
        ocr = RapidOCR()
        for shot_id, evidence_path, key in pending:
            text = ocr_subtitle(ocr, imread(evidence_path), subtitle_region)
            text = normalize_subtitle(text)
            cache[key] = text
            subtitles[shot_id] = text
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(ocr_evidence_file, evidence_path, subtitle_region): (shot_id, key)
                for shot_id, evidence_path, key in pending
            }
            for future in as_completed(future_map):
                shot_id, key = future_map[future]
                try:
                    text = future.result()
                except Exception:
                    text = ""
                text = normalize_subtitle(text)
                cache[key] = text
                subtitles[shot_id] = text

    save_ocr_cache(cache_path, cache)
    return subtitles


def sample_frames_for_shot(shot: dict[str, Any], include_full: bool) -> list[int]:
    start = int(shot.get("start_frame", 0))
    end = int(shot.get("end_frame", start))
    rep = int(shot.get("representative_frame", start))
    midpoint = start + max(0, (end - start) // 2)
    near_end = max(start, end - 2)
    # The evidence frame is usually the representative frame and has already
    # been OCR'd. Near-end and start frames catch short subtitle windows that
    # often fall just after or just before the embedded evidence frame.
    frames = [near_end, start, midpoint]
    if include_full:
        frames.append(max(start, end - 4))
    return [frame for frame in dict.fromkeys(frames) if frame != rep]


def choose_sampled_subtitle(texts: list[str]) -> str:
    cleaned = [normalize_subtitle(text) for text in texts if normalize_subtitle(text)]
    if not cleaned:
        return ""
    counts = Counter(cleaned)
    return max(cleaned, key=lambda item: (counts[item], len(item)))


def precompute_targeted_subtitles(
    shots: list[dict[str, Any]],
    video_path: Path,
    evidence_subtitles: dict[str, str],
    subtitle_region: str,
    workers: int,
    include_full: bool,
    frame_budget: int,
    cache_path: Path | None,
) -> dict[str, str]:
    targets = [
        shot
        for shot in shots
        if not evidence_subtitles.get(str(shot.get("shot_id", "")))
    ]
    if not targets:
        return {}

    def target_priority(shot: dict[str, Any]) -> tuple[int, float, float]:
        start_ms = float(shot.get("start_time_ms", 0.0))
        duration_ms = float(shot.get("duration_ms", 0.0))
        notes = str(shot.get("notes", ""))
        likely_dialog = 1 if duration_ms >= 450.0 else 0
        likely_change = 1 if any(token in notes for token in ["visual_state_change", "frame_difference_peak"]) else 0
        early_bonus = -start_ms if start_ms <= 90000.0 else -90000.0
        return (likely_dialog + likely_change, early_bonus, duration_ms)

    targets = sorted(targets, key=target_priority, reverse=True)

    shot_frames: dict[str, list[int]] = {
        str(shot.get("shot_id", "")): sample_frames_for_shot(shot, include_full)
        for shot in targets
    }
    frame_budget = max(0, int(frame_budget or 0))
    if frame_budget <= 0:
        return {}

    budgeted: dict[str, list[int]] = {}
    used: set[int] = set()
    max_samples = max((len(frames) for frames in shot_frames.values()), default=0)
    for sample_index in range(max_samples):
        for shot in targets:
            shot_id = str(shot.get("shot_id", ""))
            frames_for_shot = shot_frames.get(shot_id, [])
            if sample_index >= len(frames_for_shot):
                continue
            frame_no = frames_for_shot[sample_index]
            if frame_no in used:
                continue
            budgeted.setdefault(shot_id, []).append(frame_no)
            used.add(frame_no)
            if len(used) >= frame_budget:
                break
        if len(used) >= frame_budget:
            break
    shot_frames = budgeted
    frame_numbers = {frame_no for frames in shot_frames.values() for frame_no in frames}
    if not frame_numbers:
        return {}

    cache = load_ocr_cache(cache_path)
    frame_texts: dict[int, str] = {}
    pending_frame_numbers: set[int] = set()
    for frame_no in frame_numbers:
        key = video_frame_ocr_cache_key(video_path, frame_no, subtitle_region)
        if key in cache:
            frame_texts[frame_no] = normalize_subtitle(str(cache[key]))
        else:
            pending_frame_numbers.add(frame_no)

    frames = read_video_frames(video_path, pending_frame_numbers)
    if frames:
        workers = max(1, int(workers or 1))
        if workers == 1:
            ocr = RapidOCR()
            for frame_no, frame in frames.items():
                text = normalize_subtitle(ocr_subtitle(ocr, frame, subtitle_region))
                frame_texts[frame_no] = text
                cache[video_frame_ocr_cache_key(video_path, frame_no, subtitle_region)] = text
        else:
            with ThreadPoolExecutor(max_workers=workers) as executor:
                future_map = {
                    executor.submit(ocr_frame_image, frame, subtitle_region): frame_no
                    for frame_no, frame in frames.items()
                }
                for future in as_completed(future_map):
                    frame_no = future_map[future]
                    try:
                        text = future.result()
                    except Exception:
                        text = ""
                    text = normalize_subtitle(text)
                    frame_texts[frame_no] = text
                    cache[video_frame_ocr_cache_key(video_path, frame_no, subtitle_region)] = text
        save_ocr_cache(cache_path, cache)

    fallback: dict[str, str] = {}
    for shot_id, frames_for_shot in shot_frames.items():
        sampled_texts = [frame_texts.get(frame_no, "") for frame_no in frames_for_shot]
        text = choose_sampled_subtitle(sampled_texts)
        if text:
            fallback[shot_id] = text
    return fallback


def visual_distance(left_path: Path, right_path: Path) -> float:
    left = imread(left_path)
    right = imread(right_path)
    if left is None or right is None:
        return 1.0
    left_small = cv2.resize(left, (96, 128), interpolation=cv2.INTER_AREA)
    right_small = cv2.resize(right, (96, 128), interpolation=cv2.INTER_AREA)
    left_gray = cv2.cvtColor(left_small, cv2.COLOR_BGR2GRAY)
    right_gray = cv2.cvtColor(right_small, cv2.COLOR_BGR2GRAY)
    left_dct = cv2.dct(cv2.resize(left_gray, (32, 32)).astype(np.float32))[:8, :8]
    right_dct = cv2.dct(cv2.resize(right_gray, (32, 32)).astype(np.float32))[:8, :8]
    left_hash = left_dct > np.median(left_dct[1:, :])
    right_hash = right_dct > np.median(right_dct[1:, :])
    hash_distance = float(np.count_nonzero(left_hash != right_hash)) / 64.0
    left_hsv = cv2.cvtColor(left_small, cv2.COLOR_BGR2HSV)
    right_hsv = cv2.cvtColor(right_small, cv2.COLOR_BGR2HSV)
    left_hist = cv2.calcHist([left_hsv], [0, 1], None, [24, 24], [0, 180, 0, 256])
    right_hist = cv2.calcHist([right_hsv], [0, 1], None, [24, 24], [0, 180, 0, 256])
    cv2.normalize(left_hist, left_hist, alpha=1.0, beta=0.0, norm_type=cv2.NORM_L1)
    cv2.normalize(right_hist, right_hist, alpha=1.0, beta=0.0, norm_type=cv2.NORM_L1)
    hist_distance = float(cv2.compareHist(left_hist, right_hist, cv2.HISTCMP_BHATTACHARYYA))
    return hash_distance * 0.58 + hist_distance * 0.42


def is_fast_proof_shot(shot: dict[str, Any]) -> bool:
    notes = str(shot.get("notes", ""))
    pace = str(shot.get("pace_tag", ""))
    return (
        "manual_added_fast_single_frame_proof" in notes
        or "protected_display_text_or_color_change" in notes
        or "protected_rapid_single_frame_sequence" in notes
        or "快节奏" in pace
    )


def should_merge_same_subtitle(previous: dict[str, Any], subtitle: str, evidence_path: Path, shot: dict[str, Any]) -> bool:
    if previous.get("protected_fast_proof") or is_fast_proof_shot(shot):
        return False
    if not subtitle or subtitle != previous.get("raw_copy"):
        return False
    return visual_distance(Path(previous["evidence_frame"]), evidence_path) < 0.36


def renumber_rows(rows: list[dict[str, Any]], thumb_dir: Path) -> None:
    for index, row in enumerate(rows, start=1):
        shot_id = f"ModelShot_{index:03d}"
        row["shot_id"] = shot_id
        row["thumb_path"] = str(make_thumb(Path(row["evidence_frame"]), shot_id, thumb_dir))


def build_rows(
    video_path: Path,
    output_dir: Path,
    thumb_dir: Path,
    video_link: str = "",
    report_mode: str = "auto",
    transcript: str = "",
    timed_transcript_segments: list[dict[str, Any]] | None = None,
    no_subtitle: bool = False,
    merge_same_subtitle: bool = True,
    subtitle_region: str = "bottom",
    ocr_workers: int = 4,
    use_ocr_cache: bool = True,
    full_ocr_sampling: bool = False,
    targeted_ocr_sampling: bool = True,
    targeted_ocr_frame_budget: int = 64,
) -> tuple[list[dict[str, Any]], Path, str]:
    video_output_dir = output_dir / video_path.stem
    report_path, report = resolve_report(video_output_dir, report_mode)
    transcript_units = split_transcript(transcript)
    timed_transcript_segments = timed_transcript_segments or []
    shots = report["shots"]
    subtitle_region = choose_subtitle_region(shots, video_output_dir, subtitle_region)
    cache_path = None
    targeted_cache_path = None
    transcript_mode = bool(transcript_units or timed_transcript_segments)
    if use_ocr_cache and not transcript_mode and not no_subtitle:
        cache_path = output_dir / "shot_text_excels" / "_ocr_cache" / f"{video_path.stem}_{subtitle_region}.json"
        targeted_cache_path = (
            output_dir
            / "shot_text_excels"
            / "_ocr_cache"
            / f"{video_path.stem}_{subtitle_region}_targeted_frames.json"
        )
    evidence_subtitles = (
        {}
        if transcript_mode or no_subtitle
        else precompute_evidence_subtitles(shots, video_output_dir, subtitle_region, ocr_workers, cache_path)
    )
    targeted_subtitles = (
        {}
        if transcript_mode or no_subtitle or not targeted_ocr_sampling
        else precompute_targeted_subtitles(
            shots,
            video_path,
            evidence_subtitles,
            subtitle_region,
            ocr_workers,
            include_full=full_ocr_sampling,
            frame_budget=targeted_ocr_frame_budget,
            cache_path=targeted_cache_path,
        )
    )

    cap: cv2.VideoCapture | None = None
    fallback_ocr: RapidOCR | None = None
    rows: list[dict[str, Any]] = []
    previous_copy = ""
    for index, shot in enumerate(shots, start=1):
        evidence_path = resolve_evidence_path(video_output_dir, shot)
        timed_copy = timed_transcript_copy_for_shot(shot, timed_transcript_segments)
        manual_copy = timed_copy or transcript_copy_for_index(index, len(shots), transcript_units, previous_copy)
        ocr_subtitle = ""
        if not manual_copy and not no_subtitle:
            evidence_text = evidence_subtitles.get(str(shot.get("shot_id", "")))
            if evidence_text:
                ocr_subtitle = evidence_text
            elif targeted_subtitles.get(str(shot.get("shot_id", ""))):
                ocr_subtitle = targeted_subtitles[str(shot.get("shot_id", ""))]
            elif full_ocr_sampling or not targeted_ocr_sampling:
                if fallback_ocr is None:
                    fallback_ocr = RapidOCR()
                if cap is None:
                    cap = cv2.VideoCapture(str(video_path))
                    if not cap.isOpened():
                        raise RuntimeError(f"Cannot open video: {video_path}")
                ocr_subtitle = extract_shot_subtitle(
                    fallback_ocr,
                    cap,
                    shot,
                    subtitle_region,
                    evidence_text=evidence_text,
                    full_sampling=full_ocr_sampling,
                )
        subtitle = manual_copy or ocr_subtitle
        repeated_copy = bool(subtitle and subtitle == previous_copy)
        if merge_same_subtitle and rows and should_merge_same_subtitle(rows[-1], subtitle, evidence_path, shot):
            rows[-1]["time"] = f'{rows[-1]["time"].split(" - ")[0]} - {shot["end_time"]}'
            previous_copy = subtitle
            continue
        plan = classify_plan(index, shot, subtitle, repeated_copy, video_path.stem)
        rows.append(
            {
                "video_link": video_link if not rows else "",
                "shot_id": f"ModelShot_{index:03d}",
                "time": f'{shot["start_time"]} - {shot["end_time"]}',
                "evidence_frame": str(evidence_path),
                "thumb_path": "",
                "raw_copy": subtitle,
                "ocr_copy": ocr_subtitle,
                "manual_copy": manual_copy,
                "source_candidate_shot_id": shot.get("source_candidate_shot_id", ""),
                "protected_fast_proof": is_fast_proof_shot(shot),
                **plan,
            }
        )
        if subtitle:
            previous_copy = subtitle

    if cap is not None:
        cap.release()
    renumber_rows(rows, thumb_dir)
    return rows, report_path, subtitle_region


def value_for(row: dict[str, Any], field: str) -> str:
    mapping = {
        "视频链接": "video_link",
        "镜头": "shot_id",
        "时间": "time",
        "视频结构（画面）": "visual",
        "文案框架": "framework",
        "文案": "copy",
        "用户视角": "viewer",
        "配音": "voice",
        "音效39个": "sound",
        "视频亮点": "highlight",
    }
    return row.get(mapping[field], "")


def build_excel(rows: list[dict[str, Any]], excel_path: Path, sheet_title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    for row_index, field in enumerate(FIELDS, start=1):
        ws.cell(row=row_index, column=1, value=field)
        for column_index, shot in enumerate(rows, start=2):
            ws.cell(row=row_index, column=column_index, value=value_for(shot, field))

    label_fill = PatternFill("solid", fgColor="7A4A28")
    link_fill = PatternFill("solid", fgColor="FFF7ED")
    body_fill = PatternFill("solid", fgColor="FFFCF7")
    label_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=11)
    body_font = Font(color="1F2937", name="Microsoft YaHei", size=10)
    shot_font = Font(color="7C2D12", bold=True, name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="E8D8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=len(FIELDS), min_col=1, max_col=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
            cell.font = body_font
            cell.fill = body_fill

    for cell in ws["A"]:
        cell.fill = label_fill
        cell.font = label_font

    for column_index in range(2, len(rows) + 2):
        ws.cell(row=2, column=column_index).font = shot_font
        ws.cell(row=1, column=column_index).fill = link_fill
        ws.column_dimensions[get_column_letter(column_index)].width = 30

    ws.column_dimensions["A"].width = 18
    for row_index, height in {
        1: 48,
        2: 220,
        3: 42,
        4: 70,
        5: 64,
        6: 120,
        7: 110,
        8: 72,
        9: 54,
        10: 86,
    }.items():
        ws.row_dimensions[row_index].height = height

    for column_index, row in enumerate(rows, start=2):
        image = XLImage(row["thumb_path"])
        image.width = 135
        image.height = 240
        ws.add_image(image, f"{get_column_letter(column_index)}2")

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def validate_excel(excel_path: Path, expected_columns: int) -> dict[str, Any]:
    wb = load_workbook(excel_path)
    ws = wb.active
    labels = [ws.cell(row=i, column=1).value for i in range(1, 11)]
    assert labels == FIELDS, labels
    assert ws.max_column == expected_columns + 1, (ws.max_column, expected_columns + 1)
    assert len(ws._images) == expected_columns, (len(ws._images), expected_columns)
    copy_filled = sum(1 for col in range(2, ws.max_column + 1) if ws.cell(row=6, column=col).value)
    return {
        "excel": str(excel_path),
        "columns": expected_columns,
        "images": len(ws._images),
        "copy_filled": copy_filled,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }


def default_excel_path(output_dir: Path, video_path: Path) -> Path:
    return output_dir / "shot_text_excels" / f"{video_path.stem}_shot_text.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build horizontal shot-text Excel for one segmented short video.")
    parser.add_argument("--video-file", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=ROOT / "output" / "test")
    parser.add_argument("--excel-file", type=Path, default=None)
    parser.add_argument("--video-link", default="")
    parser.add_argument("--sheet-title", default="")
    parser.add_argument("--report-mode", choices=["auto", "calibrated", "reference", "model"], default="auto")
    parser.add_argument("--transcript-text", default="")
    parser.add_argument("--transcript-file", type=Path, default=None)
    parser.add_argument("--timed-transcript-json", type=Path, default=None)
    parser.add_argument("--no-subtitle", action="store_true")
    parser.add_argument("--disable-same-subtitle-merge", action="store_true")
    parser.add_argument(
        "--subtitle-region",
        choices=["auto", "bottom", "top", "top-bottom", "wide", "middle", "center"],
        default="auto",
    )
    parser.add_argument("--ocr-workers", type=int, default=min(4, os.cpu_count() or 1))
    parser.add_argument("--no-ocr-cache", action="store_true")
    parser.add_argument("--full-ocr-sampling", action="store_true")
    parser.add_argument("--disable-targeted-ocr-sampling", action="store_true")
    parser.add_argument("--targeted-ocr-frame-budget", type=int, default=64)
    args = parser.parse_args()

    video_path = args.video_file.resolve()
    output_dir = args.output_dir.resolve()
    excel_path = args.excel_file.resolve() if args.excel_file else default_excel_path(output_dir, video_path)
    thumb_dir = output_dir / "shot_text_excels" / "_thumbs" / video_path.stem
    thumb_dir.mkdir(parents=True, exist_ok=True)
    sheet_title = args.sheet_title or f"{video_path.stem}拆解"

    transcript = args.transcript_text
    if args.transcript_file:
        transcript = args.transcript_file.read_text(encoding="utf-8")
    timed_transcript_segments = load_timed_transcript(args.timed_transcript_json)
    rows, report_path, resolved_subtitle_region = build_rows(
        video_path,
        output_dir,
        thumb_dir,
        args.video_link,
        report_mode=args.report_mode,
        transcript=transcript,
        timed_transcript_segments=timed_transcript_segments,
        no_subtitle=args.no_subtitle,
        merge_same_subtitle=not args.disable_same_subtitle_merge,
        subtitle_region=args.subtitle_region,
        ocr_workers=args.ocr_workers,
        use_ocr_cache=not args.no_ocr_cache,
        full_ocr_sampling=args.full_ocr_sampling,
        targeted_ocr_sampling=not args.disable_targeted_ocr_sampling,
        targeted_ocr_frame_budget=args.targeted_ocr_frame_budget,
    )
    build_excel(rows, excel_path, sheet_title)
    result = validate_excel(excel_path, len(rows))
    result["video"] = str(video_path)
    result["report"] = str(report_path)
    result["subtitle_region"] = resolved_subtitle_region
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
