import json
import re
import argparse
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


ROOT = Path(r"E:\USE\codexhome\fenge")
OUTPUT_TEST = ROOT / "output" / "test"
OUTPUT_EXCEL_DIR = OUTPUT_TEST / "shot_text_excels"
OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

VIDEO_DIR = max(
    [
        path
        for path in OUTPUT_TEST.iterdir()
        if path.is_dir() and (path / "model_optimized" / "model_optimized_shot_report.json").exists()
    ],
    key=lambda path: path.stat().st_mtime,
)
REPORT_PATH = VIDEO_DIR / "model_optimized" / "model_optimized_shot_report.json"
OCR_JSON_PATH = VIDEO_DIR / "model_optimized" / "white_subtitle_ocr.json"
THUMB_DIR = OUTPUT_EXCEL_DIR / "_thumbs" / "lace_underwear"
THUMB_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = OUTPUT_EXCEL_DIR / "lace_underwear_white_subtitle_shot_text.xlsx"
DEFAULT_VIDEO_LINK = ""

FIELDS = [
    "视频链接",
    "镜头",
    "时间",
    "视频结构（画面）",
    "文案框架",
    "文案",
    "用户视角",
    "配音",
    "音效39个",
    "视频亮点",
]


def read_image(path: Path):
    data = np.fromfile(str(path), dtype=np.uint8)
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def clean_ocr_text(text: str) -> str:
    text = re.sub(r"\s+", "", text or "")
    text = text.replace("ufeel", "Ufeel").replace("UFeel", "Ufeel")
    text = re.sub(r"\b(ORRE|uimUin|Idalbs8|bICBD)\b", "", text, flags=re.I)
    text = re.sub(r"^[A-Za-z0-9]{3,}", "", text)
    text = text.replace("下半生", "下半身")
    text = text.replace("昨不穿", "咋不穿")
    text = text.replace("奶瓶", "奶皮")
    text = text.replace("买一发三", "拍一发三")
    text = text.replace("烂裤权", "烂内裤")
    text = text.replace("烂裤全", "烂内裤")
    text = re.sub(r"[?？]+", "", text)
    text = re.sub(r"^[，。、\s]+|[，。、\s]+$", "", text)
    return text


def is_picture_text(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    if re.search(r"AI生成|推荐官|黄圣依|品牌|DESIGNER|PREMIUM", compact, re.I):
        return True
    if "高腰收腹无痕内裤" in compact:
        return True
    if re.search(r"[A-Za-z0-9]{4,}.*高腰收腹无痕内裤", compact):
        return True
    if re.fullmatch(r"[A-Za-z0-9/_.-]{3,}", compact):
        return True
    return False


def extract_white_subtitle(ocr: RapidOCR, image_path: Path):
    image = read_image(image_path)
    if image is None:
        return "", [], 0.0
    h, w = image.shape[:2]
    result, _ = ocr(image)
    kept = []
    if result:
        for box, raw_text, raw_score in result:
            score = float(raw_score)
            xs = [point[0] for point in box]
            ys = [point[1] for point in box]
            x1, y1, x2, y2 = map(float, [min(xs), min(ys), max(xs), max(ys)])
            cx = (x1 + x2) / 2
            cy = (y1 + y2) / 2
            text = clean_ocr_text(str(raw_text))
            if score < 0.55:
                continue
            if cy < h * 0.60:
                continue
            if cx < w * 0.08 or cx > w * 0.92:
                continue
            if not text or text == "福" or is_picture_text(text):
                continue
            kept.append({"text": text, "score": score, "box": [round(x1), round(y1), round(x2), round(y2)]})
    kept.sort(key=lambda item: (item["box"][1], item["box"][0]))
    text = " ".join(item["text"] for item in kept).strip()
    avg = sum(item["score"] for item in kept) / len(kept) if kept else 0.0
    return text, kept, avg


def normalize_for_repeat(text: str) -> str:
    return re.sub(r"[，。！？、\s]", "", text or "")


def meaningful(row):
    text = row["white_subtitle"]
    if not text or row.get("subtitle_same_prev"):
        return False
    compact = normalize_for_repeat(text)
    if len(compact) < 4:
        return False
    if compact in {"因为", "现在买Ufeel", "穿穿穿", "继续穿", "关键是啊", "我给你看啊", "那现在呢", "它这个版型啊"}:
        return False
    return True


def has(text, pattern):
    return re.search(pattern, text or "") is not None


def visual_structure(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"我老婆只穿|翻个身"):
        return "强反差吸睛画面"
    if has(text, r"春季送福利|拍一发三|价格"):
        return "活动利益点画面"
    if has(text, r"肚子|收腹|妈妈臀|臀部|全包臀|上身"):
        return "上身效果展示"
    if has(text, r"下半身|舒服|女性健康|内裤|底档"):
        return "健康舒适痛点画面"
    if has(text, r"拆开|洗|Ufeel家的拆开"):
        return "包装/即穿证明"
    if has(text, r"镂空|蕾丝|面料|无痕|紧身裙|揉搓"):
        return "蕾丝面料细节"
    if has(text, r"健身|运动"):
        return "运动穿着场景"
    return "口播字幕画面"


def copy_framework(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"我老婆只穿|翻个身|只有两种"):
        return "反常识钩子"
    if has(text, r"春季送福利|拍一发三|价格"):
        return "活动利益点"
    if has(text, r"肚子越大|收腹效果|妈妈臀|臀部肉肉"):
        return "身材痛点+效果承诺"
    if has(text, r"下半身|舒服|女性健康"):
        return "私密舒适痛点"
    if has(text, r"拆开|洗|拆开就能穿"):
        return "方便即穿/卫生顾虑"
    if has(text, r"镂空|蕾丝|纯欲|性感|面料|无痕"):
        return "产品核心卖点"
    if has(text, r"紧身裙|运动健身"):
        return "穿搭场景验证"
    return "卖点信息补充"


def user_perspective(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"我老婆只穿|翻个身|只有两种"):
        return "被反差说法吸引，想知道为什么这条内裤值得买。"
    if has(text, r"春季送福利|拍一发三|价格"):
        return "活动利益明确，容易产生趁现在入手的想法。"
    if has(text, r"肚子|收腹|妈妈臀|臀部肉肉"):
        return "联想到自己的身材和穿搭尴尬，关注上身改善效果。"
    if has(text, r"下半身|舒服|女性健康"):
        return "把内裤和日常舒适、私密健康联系起来。"
    if has(text, r"拆开|洗|拆开就能穿"):
        return "即穿和卫生感降低购买顾虑。"
    if has(text, r"蕾丝|镂空|纯欲|性感|无痕"):
        return "既想要好看，也在意贴身穿是否舒服、不勒、不显痕。"
    if has(text, r"紧身裙|运动健身"):
        return "代入运动或紧身穿搭场景，判断是否适合自己。"
    return ""


def voiceover(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"我老婆只穿|只有两种|你说"):
        return "反问式口播，语气夸张拉停留。"
    if has(text, r"福利|拍一发三|价格"):
        return "活动强调型口播，突出便宜和限时感。"
    if has(text, r"肚子|妈妈臀|下半身|舒服"):
        return "痛点讲解型口播，语气直接。"
    if has(text, r"蕾丝|面料|无痕|全包臀"):
        return "产品讲解型口播，跟随画面点卖点。"
    return "生活化口播，节奏自然。"


def sound_effect(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"福利|拍一发三|价格"):
        return "价格 ding / 字幕 pop"
    if has(text, r"翻个身|只有两种"):
        return "转场 whoosh / 钩子 hit"
    if has(text, r"拆开|揉搓|面料|蕾丝"):
        return "布料 swish / 摩擦声"
    if has(text, r"肚子|妈妈臀|舒服"):
        return "痛点 hit / 轻提示音"
    return "轻快 BGM / 字幕 pop"


def highlight(row):
    text = row["white_subtitle"]
    if not meaningful(row):
        return ""
    if has(text, r"我老婆只穿|只有两种"):
        return "开头用反常识说法制造停留。"
    if has(text, r"福利|拍一发三|价格"):
        return "转化利益点清晰。"
    if has(text, r"肚子|收腹|妈妈臀|臀部肉肉"):
        return "身材痛点和上身效果绑定，适合重点放大。"
    if has(text, r"下半身|女性健康|舒服"):
        return "把产品从好看延伸到舒适健康。"
    if has(text, r"拆开|洗|拆开就能穿"):
        return "解决即穿和卫生顾虑。"
    if has(text, r"蕾丝|全包臀|无痕|紧身裙"):
        return "好看、包臀、无痕三类卖点集中。"
    return "承接卖点节奏。"


def make_thumb(image_path: Path, shot_id: str) -> Path:
    output = THUMB_DIR / f"{shot_id}.jpg"
    with Image.open(image_path) as img:
        img.thumbnail((190, 338))
        canvas = Image.new("RGB", (190, 338), "white")
        x = (190 - img.width) // 2
        y = (338 - img.height) // 2
        canvas.paste(img.convert("RGB"), (x, y))
        canvas.save(output, quality=88)
    return output


def build_rows():
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    ocr = RapidOCR()
    rows = []
    previous = ""
    for shot in report["shots"]:
        image_path = Path(shot["evidence_frame"])
        if not image_path.exists():
            image_path = VIDEO_DIR / "model_optimized" / "evidence" / Path(shot["evidence_frame"]).name
        subtitle, items, avg_score = extract_white_subtitle(ocr, image_path)
        repeated = bool(subtitle and normalize_for_repeat(subtitle) == previous)
        if subtitle:
            previous = normalize_for_repeat(subtitle)
        row = {
            "shot_id": shot["shot_id"],
            "start_time": shot["start_time"],
            "end_time": shot["end_time"],
            "start_frame": shot["start_frame"],
            "end_frame": shot["end_frame"],
            "action_label": shot.get("action_label", ""),
            "evidence_frame": str(image_path),
            "white_subtitle": "" if is_picture_text(subtitle) else subtitle,
            "ocr_items": items,
            "ocr_confidence": round(avg_score, 4) if items else "",
            "subtitle_same_prev": repeated,
            "thumb_path": str(make_thumb(image_path, shot["shot_id"])),
        }
        rows.append(row)
    OCR_JSON_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return rows


def value_for(row, field):
    if field == "视频链接":
        return row.get("video_link", "")
    if field == "镜头":
        return row["shot_id"]
    if field == "时间":
        return f'{row["start_time"]} - {row["end_time"]}'
    if field == "视频结构（画面）":
        return visual_structure(row)
    if field == "文案框架":
        return copy_framework(row)
    if field == "文案":
        return row["white_subtitle"] if meaningful(row) else ""
    if field == "用户视角":
        return user_perspective(row)
    if field == "配音":
        return voiceover(row)
    if field == "音效39个":
        return sound_effect(row)
    if field == "视频亮点":
        return highlight(row)
    return ""


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "蕾丝内裤拆解"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    for row_index, field in enumerate(FIELDS, start=1):
        ws.cell(row=row_index, column=1, value=field)
        for column_index, shot in enumerate(rows, start=2):
            ws.cell(row=row_index, column=column_index, value=value_for(shot, field))

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    body_font = Font(color="111827", name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=len(FIELDS), min_col=1, max_col=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
            cell.font = body_font
    for cell in ws["A"]:
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions["A"].width = 17
    for column_index in range(2, len(rows) + 2):
        ws.column_dimensions[get_column_letter(column_index)].width = 28
    ws.row_dimensions[1].height = 88
    ws.row_dimensions[2].height = 205
    for row_index in range(3, len(FIELDS) + 1):
        ws.row_dimensions[row_index].height = 72

    for column_index, row in enumerate(rows, start=2):
        image = XLImage(row["thumb_path"])
        image.width = 135
        image.height = 240
        ws.add_image(image, f"{get_column_letter(column_index)}2")

    wb.save(EXCEL_PATH)


def validate_excel(expected_columns):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    labels = [ws.cell(row=i, column=1).value for i in range(1, 11)]
    image_count = len(getattr(ws, "_images", []))
    assert labels == FIELDS, labels
    assert ws.max_column == expected_columns + 1, (ws.max_column, expected_columns + 1)
    assert image_count == expected_columns, (image_count, expected_columns)
    filled_copy = sum(1 for column in range(2, ws.max_column + 1) if ws.cell(row=5, column=column).value)
    return {"columns": expected_columns, "images": image_count, "filled_copy": filled_copy}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--video-link", default=DEFAULT_VIDEO_LINK)
    args = parser.parse_args()
    rows = build_rows()
    for index, row in enumerate(rows):
        row["video_link"] = args.video_link if index == 0 else ""
    build_excel(rows)
    stats = validate_excel(len(rows))
    print(json.dumps({"video_dir": str(VIDEO_DIR), "ocr_json": str(OCR_JSON_PATH), "excel": str(EXCEL_PATH), **stats}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
