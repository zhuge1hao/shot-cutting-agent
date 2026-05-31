import json
import re
from collections import Counter
from pathlib import Path

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


ROOT = Path(r"E:\USE\codexhome\fenge")
VIDEO_PATH = ROOT / "videos" / "test" / "5.8+文胸+都市丽人经典官方+11名.mp4"
OUTPUT_TEST = ROOT / "output" / "test"
VIDEO_DIR = OUTPUT_TEST / VIDEO_PATH.stem
CALIBRATED_REPORT_PATH = VIDEO_DIR / "model_calibrated" / "optimized_shot_report.json"
MODEL_REPORT_PATH = VIDEO_DIR / "model_optimized" / "model_optimized_shot_report.json"
REPORT_PATH = CALIBRATED_REPORT_PATH if CALIBRATED_REPORT_PATH.exists() else MODEL_REPORT_PATH
OUTPUT_EXCEL_DIR = OUTPUT_TEST / "shot_text_excels"
THUMB_DIR = OUTPUT_EXCEL_DIR / "_thumbs" / "bra_dushiliren_5_8"
EXCEL_PATH = OUTPUT_EXCEL_DIR / "bra_dushiliren_5_8_shot_text.xlsx"

OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
THUMB_DIR.mkdir(parents=True, exist_ok=True)

FIELDS = [
    "视频链接",
    "镜头",
    "时间",
    "视频结构（画面）",
    "文案框架",
    "文案",
    "用户视角",
    "配音",
    "音效39个",
    "视频亮点",
]


def imread(path: Path) -> np.ndarray | None:
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def read_video_frame(cap: cv2.VideoCapture, frame_no: int) -> np.ndarray | None:
    cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, int(frame_no)))
    ok, frame = cap.read()
    return frame if ok else None


def normalize_subtitle(text: str) -> str:
    text = re.sub(r"\s+", "", text.strip())
    text = text.replace("苹身", "半身")
    text = text.replace("善通", "普通")
    text = text.replace("普逼", "普通")
    text = text.replace("内认", "内衣")
    text = text.replace("不间", "不闷")
    text = text.replace("两杆小风扇", "两个小风扇")
    text = text.replace("文胸", "内衣") if text == "文胸" else text
    return text


def looks_like_subtitle(text: str, score: float) -> bool:
    if score < 0.62:
        return False
    text = normalize_subtitle(text)
    if len(text) < 4:
        return False
    if re.fullmatch(r"[\d:：./ -]+", text):
        return False
    if not re.search(r"[\u4e00-\u9fff]", text):
        return False
    fixed_noise = ["7:00", "jMo", "MO", "No", "mO"]
    return text not in fixed_noise


def ocr_bottom_subtitle(ocr: RapidOCR, frame: np.ndarray | None) -> str:
    if frame is None:
        return ""
    h, w = frame.shape[:2]
    # Bottom subtitle band. This avoids packaging/product text in the main picture.
    top = int(h * 0.55)
    bottom = int(h * 0.86)
    crop = frame[top:bottom, 0:w]
    result, _ = ocr(crop)
    if not result:
        return ""

    candidates: list[tuple[float, float, str]] = []
    crop_h = max(1, crop.shape[0])
    for item in result:
        box, raw_text, score = item[0], str(item[1]), float(item[2])
        text = normalize_subtitle(raw_text)
        if not looks_like_subtitle(text, score):
            continue
        ys = [point[1] for point in box]
        xs = [point[0] for point in box]
        center_y = sum(ys) / len(ys)
        box_h = max(ys) - min(ys)
        box_w = max(xs) - min(xs)
        # The subtitle is large and sits in the lower-middle of this crop.
        if center_y < crop_h * 0.38 or center_y > crop_h * 0.92:
            continue
        if box_h < 20 or box_w < w * 0.18:
            continue
        candidates.append((center_y, min(xs), text))

    if not candidates:
        return ""
    candidates.sort()
    merged: list[str] = []
    for _, _, text in candidates:
        if text not in merged:
            merged.append(text)
    return " / ".join(merged)


def extract_shot_subtitle(ocr: RapidOCR, cap: cv2.VideoCapture, shot: dict) -> str:
    start = int(shot.get("start_frame", 0))
    end = int(shot.get("end_frame", start))
    rep = int(shot.get("representative_frame", start))
    sample_frames = [
        rep,
        start,
        start + max(0, (end - start) // 2),
        max(start, end - 2),
    ]
    texts: list[str] = []
    for frame_no in dict.fromkeys(sample_frames):
        text = ocr_bottom_subtitle(ocr, read_video_frame(cap, frame_no))
        if text:
            texts.append(text)
    if not texts:
        evidence = imread(Path(str(shot.get("evidence_frame", ""))))
        text = ocr_bottom_subtitle(ocr, evidence)
        if text:
            texts.append(text)
    if not texts:
        return ""

    counts = Counter(texts)
    # Prefer the most repeated subtitle; tie-break by longer text.
    return max(counts, key=lambda item: (counts[item], len(item)))


def make_thumb(image_path: Path, shot_id: str) -> Path:
    output = THUMB_DIR / f"{shot_id}.jpg"
    with Image.open(image_path) as img:
        img.thumbnail((190, 338), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (190, 338), "white")
        x = (190 - img.width) // 2
        y = (338 - img.height) // 2
        canvas.paste(img.convert("RGB"), (x, y))
        canvas.save(output, quality=88)
    return output


def classify_plan(index: int, shot: dict, copy: str, repeated_copy: bool) -> dict:
    text = copy
    notes = str(shot.get("notes", ""))
    pace = str(shot.get("pace_tag", ""))
    start_ms = float(shot.get("start_time_ms", 0.0))

    visual = "模特/产品画面承接"
    framework = ""
    viewer = ""

    if index <= 3:
        visual = "开场舒适痛点"
        framework = "舒适需求开场"
        viewer = "先用上半身舒适把目标用户带入，建立继续看的理由。"
    elif "普通内衣" in text or "不要再穿" in text:
        visual = "普通内衣痛点对比"
        framework = "痛点提醒"
        viewer = "用户会联想到普通内衣勒、空杯、跑杯或不舒服的问题。"
    elif "提拉内衣" in text or "都市丽人" in text:
        visual = "提拉内衣正式露出"
        framework = "产品给到解决方案"
        viewer = "明确产品类型和品牌，开始判断是否适合自己。"
    elif any(word in text for word in ["一口气", "4件", "两件"]):
        visual = "多件囤货/复购证明"
        framework = "数量价值证明"
        viewer = "看到一次买多件，会觉得产品有换新和囤货价值。"
    elif "25岁" in text or "上身" in text:
        visual = "模特上身展示"
        framework = "上身效果证明"
        viewer = "通过模特上身判断版型、显瘦和日常穿搭效果。"
    elif any(word in text for word in ["旧内衣", "全扔"]):
        visual = "旧内衣淘汰痛点"
        framework = "换新理由强化"
        viewer = "旧内衣被淘汰，强化用户也该换新的心理。"
    elif any(word in text for word in ["巨隐形", "蜜桃胸", "蚂蚁腰"]):
        visual = "隐形显身材效果"
        framework = "身材结果卖点"
        viewer = "用户会关注穿上后是否显胸型、显腰细、外穿不露痕。"
    elif any(word in text for word in ["春夏", "穿啥都好看", "白t", "罩衫"]):
        visual = "春夏穿搭场景"
        framework = "穿搭场景验证"
        viewer = "把产品放进白T、罩衫、薄衣服等真实场景里判断实用性。"
    elif any(word in text for word in ["娜扎", "同款"]):
        visual = "明星同款背书"
        framework = "种草信任背书"
        viewer = "明星同款会降低陌生感，提高继续看的兴趣。"
    elif any(word in text for word in ["27年", "4500", "6300万", "国民", "门店", "女性选择", "品质"]):
        visual = "品牌实力背书"
        framework = "品牌信任建立"
        viewer = "通过品牌年限、门店和用户规模判断是否可靠。"
    elif any(word in text for word in ["喜马拉雅", "太极磁", "四道工序", "轻氧SPA", "软又滑"]):
        visual = "面料科技/触感证明"
        framework = "材料卖点拆解"
        viewer = "从面料来源、工艺和触感判断是否真的舒服。"
    elif any(word in text for word in ["大胸", "小胸", "胸型"]):
        visual = "不同胸型适配"
        framework = "适用人群扩展"
        viewer = "大胸小胸用户都能找到对应购买理由。"
    elif any(word in text for word in ["轻盈", "透气", "舒适", "裸感"]):
        visual = "轻盈舒适证明"
        framework = "舒适卖点"
        viewer = "关注内衣是否闷、勒、厚，以及春夏穿着是否舒服。"
    elif any(word in text for word in ["不闷", "透气网孔", "凉感", "小风扇", "双重透气"]):
        visual = "透气凉感证明"
        framework = "夏季舒适卖点"
        viewer = "夏天用户最担心闷热出汗，这里用网孔和凉感降低顾虑。"
    elif any(word in text for word in ["杯垫", "防凸", "跑杯"]):
        visual = "杯垫/防凸点证明"
        framework = "结构痛点解决"
        viewer = "解决杯垫移位、凸点和清洗后跑杯的顾虑。"
    elif any(word in text for word in ["副乳", "侧收", "支撑", "腋下"]):
        visual = "侧收软支撑证明"
        framework = "支撑结构拆解"
        viewer = "关注副乳收拢、侧边稳定和无钢圈支撑是否够用。"
    elif any(word in text for word in ["美背", "交叉", "露背", "肩带", "深V"]):
        visual = "美背/肩带穿法展示"
        framework = "穿搭场景验证"
        viewer = "判断吊带、露背、薄上衣等场景能不能搭配。"
    elif any(word in text for word in ["法国设计师", "尴尬留印", "留印"]):
        visual = "无痕外穿证明"
        framework = "尴尬痛点解决"
        viewer = "关注穿白T、罩衫时是否会透出内衣痕迹。"
    elif any(word in text for word in ["10A", "抗菌", "300次", "放心"]):
        visual = "抗菌耐洗证明"
        framework = "品质安全收口"
        viewer = "抗菌和水洗次数用于打消卫生、耐用和安全顾虑。"
    elif any(word in text for word in ["活动", "价格", "拍", "下单", "闭眼"]):
        visual = "活动转化收口"
        framework = "价格活动促单"
        viewer = "卖点建立后，价格和活动推动用户下单。"
    elif "快节奏" in pace or "rapid" in notes:
        visual = "快速产品证明单帧"
        framework = "快切卖点补充"
        viewer = "通过短促单帧快速确认产品细节和穿搭效果。"
    elif start_ms < 12000:
        visual = "开场产品/上身切换"
        framework = "开场钩子承接"
        viewer = "用连续快切把痛点、产品和上身状态串起来。"

    if repeated_copy:
        framework = ""
        viewer = ""

    return {
        "visual": visual,
        "framework": framework,
        "copy": "" if repeated_copy else text,
        "viewer": viewer,
        "voice": "生活化种草口播，语速轻快，重点词稍加强。" if framework else "",
        "sound": "轻快BGM / 字幕pop / 细节ding" if framework else "",
        "highlight": "该镜头承担卖点证明或转场承接作用。" if framework else "",
    }


def resolve_evidence_path(shot: dict) -> Path:
    path = Path(str(shot.get("evidence_frame", "")))
    if path.exists():
        return path
    return VIDEO_DIR / "model_optimized" / "evidence" / path.name


def build_rows(video_link: str = "") -> list[dict]:
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    cap = cv2.VideoCapture(str(VIDEO_PATH))
    if not cap.isOpened():
        raise RuntimeError(f"Cannot open video: {VIDEO_PATH}")
    ocr = RapidOCR()

    rows = []
    previous_copy = ""
    for index, shot in enumerate(report["shots"], start=1):
        evidence_path = resolve_evidence_path(shot)
        subtitle = extract_shot_subtitle(ocr, cap, shot)
        repeated_copy = bool(subtitle and subtitle == previous_copy)
        plan = classify_plan(index, shot, subtitle, repeated_copy)
        thumb = make_thumb(evidence_path, f"ModelShot_{index:03d}")
        rows.append(
            {
                "video_link": video_link if index == 1 else "",
                "shot_id": f"ModelShot_{index:03d}",
                "time": f'{shot["start_time"]} - {shot["end_time"]}',
                "evidence_frame": str(evidence_path),
                "thumb_path": str(thumb),
                **plan,
            }
        )
        if subtitle:
            previous_copy = subtitle

    cap.release()
    return rows


def value_for(row: dict, field: str) -> str:
    mapping = {
        "视频链接": "video_link",
        "镜头": "shot_id",
        "时间": "time",
        "视频结构（画面）": "visual",
        "文案框架": "framework",
        "文案": "copy",
        "用户视角": "viewer",
        "配音": "voice",
        "音效39个": "sound",
        "视频亮点": "highlight",
    }
    return row.get(mapping[field], "")


def build_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "都市丽人提拉内衣拆解"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    for row_index, field in enumerate(FIELDS, start=1):
        ws.cell(row=row_index, column=1, value=field)
        for column_index, shot in enumerate(rows, start=2):
            ws.cell(row=row_index, column=column_index, value=value_for(shot, field))

    label_fill = PatternFill("solid", fgColor="7A4A28")
    link_fill = PatternFill("solid", fgColor="FFF7ED")
    body_fill = PatternFill("solid", fgColor="FFFCF7")
    label_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=11)
    body_font = Font(color="1F2937", name="Microsoft YaHei", size=10)
    shot_font = Font(color="7C2D12", bold=True, name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="E8D8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=len(FIELDS), min_col=1, max_col=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
            cell.font = body_font
            cell.fill = body_fill

    for cell in ws["A"]:
        cell.fill = label_fill
        cell.font = label_font

    for column_index in range(2, len(rows) + 2):
        ws.cell(row=2, column=column_index).font = shot_font
        ws.cell(row=1, column=column_index).fill = link_fill
        ws.column_dimensions[get_column_letter(column_index)].width = 30

    ws.column_dimensions["A"].width = 18
    row_heights = {
        1: 48,
        2: 220,
        3: 42,
        4: 70,
        5: 64,
        6: 120,
        7: 110,
        8: 72,
        9: 54,
        10: 86,
    }
    for row_index, height in row_heights.items():
        ws.row_dimensions[row_index].height = height

    for column_index, row in enumerate(rows, start=2):
        image = XLImage(row["thumb_path"])
        image.width = 135
        image.height = 240
        ws.add_image(image, f"{get_column_letter(column_index)}2")

    wb.save(EXCEL_PATH)


def validate_excel(expected_columns: int) -> dict:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    labels = [ws.cell(row=i, column=1).value for i in range(1, 11)]
    assert labels == FIELDS, labels
    assert ws.max_column == expected_columns + 1, (ws.max_column, expected_columns + 1)
    assert len(ws._images) == expected_columns, (len(ws._images), expected_columns)
    copy_filled = sum(1 for col in range(2, ws.max_column + 1) if ws.cell(row=6, column=col).value)
    return {
        "excel": str(EXCEL_PATH),
        "video_dir": str(VIDEO_DIR),
        "columns": expected_columns,
        "images": len(ws._images),
        "copy_filled": copy_filled,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }


def main() -> None:
    rows = build_rows()
    build_excel(rows)
    print(json.dumps(validate_excel(len(rows)), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
