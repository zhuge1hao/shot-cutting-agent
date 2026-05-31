import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


ROOT = Path(r"E:\USE\codexhome\fenge")
OUTPUT_TEST = ROOT / "output" / "test"
OUTPUT_EXCEL_DIR = OUTPUT_TEST / "shot_text_excels"
OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

VIDEO_DIR = max(
    [
        path
        for path in OUTPUT_TEST.iterdir()
        if path.is_dir()
        and path.name.startswith("5.8+")
        and (path / "model_optimized" / "model_optimized_shot_report.json").exists()
    ],
    key=lambda path: path.stat().st_mtime,
)
REFERENCE_REPORT_PATH = VIDEO_DIR / "reference_optimized" / "optimized_shot_report.json"
MODEL_REPORT_PATH = VIDEO_DIR / "model_optimized" / "model_optimized_shot_report.json"
REPORT_PATH = REFERENCE_REPORT_PATH if REFERENCE_REPORT_PATH.exists() else MODEL_REPORT_PATH
THUMB_DIR = OUTPUT_EXCEL_DIR / "_thumbs" / "bra_youkeshushu_5_8"
THUMB_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = OUTPUT_EXCEL_DIR / "bra_youkeshushu_5_8_shot_text.xlsx"

FIELDS = [
    "视频链接",
    "镜头",
    "时间",
    "视频结构（画面）",
    "文案框架",
    "文案",
    "用户视角",
    "配音",
    "音效39个",
    "视频亮点",
]

SHOT_PLAN = [
    {
        "visual": "情侣同框关系钩子",
        "framework": "关系利益开场",
        "copy": "为了两个人的和谐，早上起来先换上普拉提老师推荐的提拉内衣。",
        "viewer": "从亲密关系场景进入，用户会先被“为了两个人的和谐”这句反常识说法拉住。",
        "voice": "生活化口播，语气带一点调侃和私密分享感。",
        "sound": "轻快 BGM / 开场 hit",
        "highlight": "用关系话题制造停留，再自然转到内衣解决方案。",
    },
    {
        "visual": "晨起上身/身材状态",
        "framework": "早晨换穿场景",
        "copy": "早上起来先换上普拉提老师推荐的提拉内衣。",
        "viewer": "代入早上换衣场景，关注内衣是否能让身形更挺、更利落。",
        "voice": "第一人称种草口播。",
        "sound": "衣物 swish / 字幕 pop",
        "highlight": "把产品使用动作放到日常起床场景里，可信度更强。",
    },
    {
        "visual": "快递盒/开箱前置",
        "framework": "囤货信任证明",
        "copy": "这是二胎后的我果断把旧内衣全扔了，一口气买了4件。",
        "viewer": "看到“二胎后”和一次买 4 件，会判断这是复购型、换新型需求。",
        "voice": "转折式口播，强调果断和数量。",
        "sound": "开箱 pop / 数量 ding",
        "highlight": "用“扔旧内衣+买4件”建立强需求和购买决心。",
    },
    {
        "visual": "产品包装+文胸正面",
        "framework": "产品正式露出",
        "copy": "一口气买了4件，有棵树提拉内衣。",
        "viewer": "明确产品品类和品牌，开始看设计细节。",
        "voice": "品牌名清晰重读。",
        "sound": "品牌露出 ding",
        "highlight": "包装和产品同框，完成从故事到商品的切换。",
    },
    {
        "visual": "模特背面穿着展示",
        "framework": "上身效果预告",
        "copy": "有棵树提拉内衣，我的妈呀，能把提拉内衣细肩带深V美背还有法式三角杯做在同一件上的，一定是个天才设计师。",
        "viewer": "先看到背部效果，关注细肩带、深V和美背是否好搭衣服。",
        "voice": "惊喜感口播，语速略快。",
        "sound": "惊喜 wow / 转场 whoosh",
        "highlight": "用上身背面先把“美背”卖点可视化。",
    },
    {
        "visual": "旧内衣丢弃画面",
        "framework": "旧品淘汰痛点",
        "copy": "二胎后的我果断把旧内衣全扔了。",
        "viewer": "理解为旧内衣不合身、不好看、不舒服，需要整体换新。",
        "voice": "果断式口播，强调扔掉旧品。",
        "sound": "丢弃 hit / 短促 whoosh",
        "highlight": "用“扔掉”动作强化换新理由。",
    },
    {
        "visual": "4件装平铺展示",
        "framework": "数量价值展示",
        "copy": "一口气买了4件。",
        "viewer": "看到多件装，会联想到换洗、囤货和价格划算。",
        "voice": "短句强调数量。",
        "sound": "数量 ding / 字幕 pop",
        "highlight": "快速建立套装感和囤货理由。",
    },
    {
        "visual": "主播手持上身展示",
        "framework": "情绪承接",
        "copy": "我的妈呀。",
        "viewer": "被主播反应带动，期待接下来讲具体设计。",
        "voice": "惊叹语气，制造节奏点。",
        "sound": "wow / 停顿 hit",
        "highlight": "用情绪短句承接产品卖点。",
    },
    {
        "visual": "黑内搭叠穿展示",
        "framework": "多卖点合一",
        "copy": "能把提拉内衣细肩带深V美背还有法式三角杯做在同一件上的，一定是个天才设计师。",
        "viewer": "开始把产品理解成一件多功能内衣，而不是普通文胸。",
        "voice": "卖点连读，突出“同一件”。",
        "sound": "卖点连击 pop",
        "highlight": "细肩带、深V、美背、三角杯集中抛出。",
    },
    {
        "visual": "浅色针织上身正面",
        "framework": "版型高级感",
        "copy": "这不就是很多超模都在穿的超薄裸感三角杯吗？",
        "viewer": "用“超模同款感”联想到显瘦、松弛和高级穿搭。",
        "voice": "反问式口播，带种草感。",
        "sound": "反问 hit / 镜头 pop",
        "highlight": "把功能内衣提升到穿搭审美层面。",
    },
    {
        "visual": "整套穿搭正面展示",
        "framework": "春夏百搭定位",
        "copy": "显瘦高级又自带松弛感，简直就是咱们春夏穿衣的百搭神器。",
        "viewer": "关心春夏薄衣服内搭是否显瘦、自然、不尴尬。",
        "voice": "轻松推荐语气。",
        "sound": "轻快 BGM / 展示 swish",
        "highlight": "把产品定位为春夏穿搭基础单品。",
    },
    {
        "visual": "多色产品/杯型平铺",
        "framework": "杯型卖点展示",
        "copy": "超薄的法式三角杯垫，精准防凸点的同时呢，又能完美贴合各种胸型。",
        "viewer": "重点关注杯垫薄不薄、防凸点、适不适合自己的胸型。",
        "voice": "讲解型口播，语速放慢。",
        "sound": "细节 ding / 布料 swish",
        "highlight": "杯垫、防凸点、贴合胸型是核心转化点。",
    },
    {
        "visual": "黑色上身显瘦效果",
        "framework": "不同身材适配",
        "copy": "大胸穿的精致显瘦，小胸穿的又高级又很斜。",
        "viewer": "无论大胸小胸，都能找到自己的购买理由。",
        "voice": "对比式口播。",
        "sound": "对比 pop / 轻快 BGM",
        "highlight": "覆盖不同胸型顾虑，扩大适用人群。",
    },
    {
        "visual": "背面/侧面美背展示",
        "framework": "上身线条验证",
        "copy": "大胸穿的精致显瘦，小胸穿的又高级又很斜。",
        "viewer": "看侧背线条是否干净，判断显瘦效果是否真实。",
        "voice": "继续解释穿着效果。",
        "sound": "转身 whoosh",
        "highlight": "用身体转动补充前一条卖点的视觉证据。",
    },
    {
        "visual": "胸前近景/贴合效果",
        "framework": "防凸点贴合证明",
        "copy": "超薄的法式三角杯垫，精准防凸点的同时呢，又能完美贴合各种胸型。",
        "viewer": "从近景判断杯垫是否服帖、是否会显尴尬印子。",
        "voice": "细节讲解型。",
        "sound": "细节放大 ding",
        "highlight": "近景增强杯垫贴合的可信度。",
    },
    {
        "visual": "侧面胸型近景",
        "framework": "胸型包容验证",
        "copy": "又能完美贴合各种胸型。",
        "viewer": "关注侧面是否空杯、压胸或外扩。",
        "voice": "补充说明式口播。",
        "sound": "细节 pop",
        "highlight": "侧面镜头补充不同胸型的适配证据。",
    },
    {
        "visual": "杯垫结构展示",
        "framework": "固定杯垫卖点",
        "copy": "还是固定式杯垫，随便你怎么清洗它都很难跑杯。",
        "viewer": "解决洗完杯垫移位、跑杯、变形的使用痛点。",
        "voice": "痛点解决型口播。",
        "sound": "杯垫 tap / 重点 ding",
        "highlight": "固定式杯垫是很强的日常维护卖点。",
    },
    {
        "visual": "指向杯垫/纹理细节",
        "framework": "结构细节证明",
        "copy": "随便你怎么清洗它都很难跑杯。",
        "viewer": "通过手指指向细节，确认不是单纯口播承诺。",
        "voice": "解释型口播。",
        "sound": "指向 tap / 字幕 pop",
        "highlight": "手部指示让结构卖点更具体。",
    },
    {
        "visual": "杯垫面料触摸",
        "framework": "舒适触感证明",
        "copy": "一位以为多层果冻条软支撑一直延伸到腋下，再到肩带的位置。",
        "viewer": "开始关注支撑是否软、是否会勒、是否能收副乳。",
        "voice": "结构讲解，突出“软支撑”。",
        "sound": "布料摩擦 swish",
        "highlight": "从杯垫过渡到支撑结构。",
    },
    {
        "visual": "穿搭上身活动展示",
        "framework": "软支撑场景验证",
        "copy": "能收住侧面副乳赘的同时呢，随便你怎么抬脚运动，它都不会上穿跑杯。",
        "viewer": "关心副乳、抬手运动、跑杯这些真实穿着问题。",
        "voice": "场景验证型口播。",
        "sound": "动作 whoosh / 重点 hit",
        "highlight": "把支撑卖点放到运动和日常动作中验证。",
    },
    {
        "visual": "边缘/果冻条近景",
        "framework": "软支撑结构拆解",
        "copy": "多层果冻条软支撑一直延伸到腋下，再到肩带的位置。",
        "viewer": "看到支撑路径，理解它为什么能收副乳、稳住杯型。",
        "voice": "专业讲解型。",
        "sound": "细节 ding / 布料 swish",
        "highlight": "结构路径可视化，增强专业感。",
    },
    {
        "visual": "白衬衫/外搭场景",
        "framework": "抬手不跑杯验证",
        "copy": "随便你怎么抬脚运动，它都不会上穿跑杯。",
        "viewer": "代入日常活动和穿外套场景，判断稳定性。",
        "voice": "轻松演示口播。",
        "sound": "动作 swish",
        "highlight": "用穿搭动作证明不会乱跑。",
    },
    {
        "visual": "小格纹杯面/深V网纱",
        "framework": "颜值设计卖点",
        "copy": "精髓还是它这个经典的小格纹搭配这个深V纯运小网纱。",
        "viewer": "从功能转到审美，关注好看、精致、可外露搭配。",
        "voice": "审美种草语气。",
        "sound": "精致 ding / 细节 pop",
        "highlight": "小格纹和深V网纱让产品从实用变成穿搭单品。",
    },
    {
        "visual": "后背交叉结构展示",
        "framework": "穿法灵活证明",
        "copy": "后背是可以交叉也可以正常穿的。",
        "viewer": "看到肩带变化，理解它能适配不同衣服。",
        "voice": "说明式口播。",
        "sound": "肩带 swish / 切换 ding",
        "highlight": "交叉/正常两穿法扩大搭配场景。",
    },
    {
        "visual": "露背穿搭背面",
        "framework": "露背搭配场景",
        "copy": "你像咱们春夏搭配个什么大露背的衣服，依旧是浅。",
        "viewer": "关心露背、吊带、薄外套时背后是否好看。",
        "voice": "穿搭建议式口播。",
        "sound": "转身 whoosh",
        "highlight": "用露背场景展示美背价值。",
    },
    {
        "visual": "背面穿法二次展示",
        "framework": "美背效果强化",
        "copy": "后背是可以交叉也可以正常穿的。",
        "viewer": "比较不同穿法的背部效果，降低搭配顾虑。",
        "voice": "补充说明。",
        "sound": "切换 pop",
        "highlight": "连续背面镜头强化可搭配性。",
    },
    {
        "visual": "杯边/面料近景",
        "framework": "无钢圈舒适卖点",
        "copy": "前美件厚美背一整件，它都是无钢圈的设计，上身真的没有一点束缚感。",
        "viewer": "关注无钢圈是否还能支撑、是否舒服不勒。",
        "voice": "舒适体验型口播。",
        "sound": "柔软 swish / 重点 ding",
        "highlight": "把颜值卖点收回到舒适穿着体验。",
    },
    {
        "visual": "面料拉伸/无束缚证明",
        "framework": "舒适度验证",
        "copy": "上身真的没有一点束缚感。",
        "viewer": "通过拉伸动作判断面料弹性和无压迫感。",
        "voice": "体验背书式口播。",
        "sound": "拉伸 swish",
        "highlight": "用手部拉伸让“无束缚”更可信。",
    },
    {
        "visual": "活动价格/主播收口",
        "framework": "活动转化收口",
        "copy": "关键好关键，他们家现在开春有活动，两件到手才这个价格。",
        "viewer": "前面卖点建立后，价格活动会推动下单。",
        "voice": "活动强调型口播，语气更急促。",
        "sound": "价格 ding / 下单 pop",
        "highlight": "用活动价格完成转化推动。",
    },
    {
        "visual": "多件产品收尾展示",
        "framework": "人群号召收尾",
        "copy": "跟我一样，春夏穿衣，喜欢这种显瘦高级感觉的你们就直接闭眼抽。",
        "viewer": "明确自己属于春夏穿搭、显瘦高级需求人群，形成购买理由。",
        "voice": "号召式口播，结尾干脆。",
        "sound": "收尾 hit / 轻快 BGM",
        "highlight": "用“显瘦高级”总结核心价值，并以闭眼冲收口。",
    },
]


def make_thumb(image_path: Path, shot_id: str) -> Path:
    output = THUMB_DIR / f"{shot_id}.jpg"
    with Image.open(image_path) as img:
        img.thumbnail((190, 338), Image.Resampling.LANCZOS)
        canvas = Image.new("RGB", (190, 338), "white")
        x = (190 - img.width) // 2
        y = (338 - img.height) // 2
        canvas.paste(img.convert("RGB"), (x, y))
        canvas.save(output, quality=88)
    return output


TIMED_SECTIONS = [
    (0, 1600, "关系利益开场", "情侣同框关系钩子", "为了两个人的和谐，早上起来先换上普拉提老师推荐的提拉内衣。", "从亲密关系场景进入，用户会先被“为了两个人的和谐”这句反常识说法拉住。"),
    (1600, 3600, "早晨换穿场景", "晨起上身/身材状态", "早上起来先换上普拉提老师推荐的提拉内衣。", "代入早上换衣场景，关注内衣是否能让身形更挺、更利落。"),
    (3600, 6600, "换新痛点+囤货证明", "快递盒/产品开箱/旧内衣淘汰", "这是二胎后的我果断把旧内衣全扔了，一口气买了4件，有棵树提拉内衣。", "看到“二胎后”和一次买 4 件，会判断这是复购型、换新型需求。"),
    (6600, 11200, "产品正式露出", "4件装/产品平铺/多件装展示", "一口气买了4件，有棵树提拉内衣。", "明确产品品类和品牌，开始看设计细节。"),
    (11200, 15100, "多卖点合一", "细肩带/深V美背/法式三角杯展示", "我的妈呀，能把提拉内衣细肩带深V美背还有法式三角杯做在同一件上的，一定是个天才设计师。", "开始把产品理解成一件多功能内衣，而不是普通文胸。"),
    (15100, 20500, "版型高级感", "超模感上身/穿搭参考", "这不就是很多超模都在穿的超薄裸感三角杯吗？", "用“超模同款感”联想到显瘦、松弛和高级穿搭。"),
    (20500, 24500, "春夏百搭定位", "显瘦高级/松弛感穿搭", "显瘦高级又自带松弛感，简直就是咱们春夏穿衣的百搭神器。", "关心春夏薄衣服内搭是否显瘦、自然、不尴尬。"),
    (24500, 31200, "杯垫核心卖点", "法式三角杯垫/防凸点/贴合胸型", "超薄的法式三角杯垫，精准防凸点的同时呢，又能完美贴合各种胸型。", "重点关注杯垫薄不薄、防凸点、适不适合自己的胸型。"),
    (31200, 37000, "不同身材适配", "大胸小胸上身对比", "大胸穿的精致显瘦，小胸穿的又高级又很又斜。", "无论大胸小胸，都能找到自己的购买理由。"),
    (37000, 43200, "固定杯垫卖点", "杯垫结构/清洗不跑杯", "还是固定式杯垫，随便你怎么清洗它都很难跑杯。", "解决洗完杯垫移位、跑杯、变形的使用痛点。"),
    (43200, 49200, "软支撑结构拆解", "果冻条/腋下/肩带支撑路径", "一位以为多层果冻条软支撑一直延伸到腋下，再到肩带的位置。", "看到支撑路径，理解它为什么能收副乳、稳住杯型。"),
    (49200, 55000, "运动稳定验证", "副乳收拢/抬手运动/不上窜跑杯", "能收住侧面副乳赘的同时呢，随便你怎么抬脚运动，它都不会上穿跑杯。", "关心副乳、抬手运动、跑杯这些真实穿着问题。"),
    (55000, 61200, "颜值设计卖点", "小格纹/深V网纱/法式浪漫", "但不过，精髓还是它这个经典的小格纹搭配这个深V纯运小网纱。你再穿个什么大零克的衣服，那真正的就是把法式精致浪漫展型的淋漓精致了。", "从功能转到审美，关注好看、精致、可外露搭配。"),
    (61200, 66000, "穿法灵活证明", "后背交叉/正常穿两种肩带", "后背是可以交叉也可以正常穿的。", "看到肩带变化，理解它能适配不同衣服。"),
    (66000, 69000, "露背搭配场景", "大露背衣服/美背效果", "你像咱们春夏搭配个什么大露背的衣服，依旧是浅。", "关心露背、吊带、薄外套时背后是否好看。"),
    (69000, 74000, "舒适+活动收口", "无钢圈舒适/活动价格/购买号召", "前美件厚美背一整件，它都是无钢圈的设计，上身真的没有一点束缚感。关键好关键，他们家现在开春有活动，两件到手才这个价格，跟我一样，春夏穿衣，喜欢这种显瘦高级感觉的你们就直接闭眼抽。", "前面卖点建立后，用无束缚和活动价格推动下单。"),
]


SPECIAL_VISUAL = {
    "Shot_028": "产品细节证据：有钢圈/无钢圈对比前的内衣结构",
    "Shot_033": "美背单帧：背面露肤和肩带位置",
    "Shot_035": "法式三角杯单帧：杯型完整露出",
    "Shot_040": "快速穿搭参考单帧：街拍黑外套",
    "Shot_041": "快速穿搭参考单帧：日常外套搭配",
    "Shot_042": "快速穿搭参考单帧：吊带裙搭配",
    "Shot_046": "上身松弛感证明：正面穿着状态",
    "Shot_055": "法式三角杯结构近景：杯面和边缘",
    "Shot_087": "穿搭反馈镜头：白色上身显瘦效果",
}


def plan_for_shot(shot: dict, previous_section: str | None) -> tuple[dict, str]:
    start_ms = float(shot.get("start_time_ms", 0.0))
    source_id = shot.get("source_candidate_shot_id", "")
    section = TIMED_SECTIONS[-1]
    for candidate in TIMED_SECTIONS:
        if candidate[0] <= start_ms < candidate[1]:
            section = candidate
            break
    _, _, framework, visual, copy, viewer = section
    first_in_section = framework != previous_section
    if source_id in SPECIAL_VISUAL:
        visual = SPECIAL_VISUAL[source_id]
        framework = "关键证据帧"
        first_in_section = True
    return (
        {
            "visual": visual,
            "framework": framework if first_in_section else "",
            "copy": copy if first_in_section else "",
            "viewer": viewer if first_in_section else "",
            "voice": "生活化种草口播，跟随画面强调对应卖点。" if first_in_section else "",
            "sound": "轻快 BGM / 字幕 pop / 细节 ding" if first_in_section else "",
            "highlight": "该镜头承担产品证明或卖点转场作用，适合作为拆解节点。" if first_in_section else "",
        },
        framework,
    )


def resolve_evidence_path(video_dir: Path, evidence_frame: str) -> Path:
    path = Path(evidence_frame)
    if path.exists():
        return path
    return video_dir / "model_optimized" / "evidence" / path.name


def build_rows(video_link: str = "") -> list[dict]:
    report = json.loads(REPORT_PATH.read_text(encoding="utf-8"))
    rows = []
    previous_section = None
    for index, shot in enumerate(report["shots"]):
        plan, previous_section = plan_for_shot(shot, previous_section)
        image_path = resolve_evidence_path(VIDEO_DIR, shot["evidence_frame"])
        rows.append(
            {
                "video_link": video_link if index == 0 else "",
                "shot_id": f"ModelShot_{index + 1:03d}",
                "source_shot_id": shot["shot_id"],
                "time": f'{shot["start_time"]} - {shot["end_time"]}',
                "start_frame": shot["start_frame"],
                "end_frame": shot["end_frame"],
                "action_label": shot.get("action_label", ""),
                "evidence_frame": str(image_path),
                "thumb_path": str(make_thumb(image_path, shot["shot_id"])),
                **plan,
            }
        )
    return rows


def value_for(row: dict, field: str) -> str:
    if field == "视频链接":
        return row["video_link"]
    if field == "镜头":
        return row["shot_id"]
    if field == "时间":
        return row["time"]
    if field == "视频结构（画面）":
        return row.get("visual", "")
    if field == "文案框架":
        return row.get("framework", "")
    if field == "文案":
        return row.get("copy", "")
    if field == "用户视角":
        return row.get("viewer", "")
    if field == "配音":
        return row.get("voice", "")
    if field == "音效39个":
        return row.get("sound", "")
    if field == "视频亮点":
        return row.get("highlight", "")
    return ""


def build_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "提拉内衣拆解"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    for row_index, field in enumerate(FIELDS, start=1):
        ws.cell(row=row_index, column=1, value=field)
        for column_index, shot in enumerate(rows, start=2):
            ws.cell(row=row_index, column=column_index, value=value_for(shot, field))

    label_fill = PatternFill("solid", fgColor="8B5A2B")
    link_fill = PatternFill("solid", fgColor="FFF7ED")
    body_fill = PatternFill("solid", fgColor="FFFBF5")
    label_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei", size=11)
    body_font = Font(color="1F2937", name="Microsoft YaHei", size=10)
    shot_font = Font(color="7C2D12", bold=True, name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="E5D3BF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=len(FIELDS), min_col=1, max_col=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
            cell.font = body_font
            cell.fill = body_fill

    for cell in ws["A"]:
        cell.fill = label_fill
        cell.font = label_font

    for column_index in range(2, len(rows) + 2):
        ws.cell(row=2, column=column_index).font = shot_font
        ws.cell(row=1, column=column_index).fill = link_fill

    ws.column_dimensions["A"].width = 18
    for column_index in range(2, len(rows) + 2):
        ws.column_dimensions[get_column_letter(column_index)].width = 30

    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 220
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[4].height = 70
    ws.row_dimensions[5].height = 64
    ws.row_dimensions[6].height = 130
    ws.row_dimensions[7].height = 120
    ws.row_dimensions[8].height = 76
    ws.row_dimensions[9].height = 56
    ws.row_dimensions[10].height = 96

    for column_index, row in enumerate(rows, start=2):
        image = XLImage(row["thumb_path"])
        image.width = 135
        image.height = 240
        ws.add_image(image, f"{get_column_letter(column_index)}2")

    wb.save(EXCEL_PATH)


def validate_excel(expected_columns: int) -> dict:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    labels = [ws.cell(row=i, column=1).value for i in range(1, 11)]
    image_count = len(getattr(ws, "_images", []))
    assert labels == FIELDS, labels
    assert ws.max_column == expected_columns + 1, (ws.max_column, expected_columns + 1)
    assert image_count == expected_columns, (image_count, expected_columns)
    return {
        "excel": str(EXCEL_PATH),
        "video_dir": str(VIDEO_DIR),
        "columns": expected_columns,
        "images": image_count,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }


def main() -> None:
    rows = build_rows(video_link="")
    build_excel(rows)
    print(json.dumps(validate_excel(len(rows)), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
